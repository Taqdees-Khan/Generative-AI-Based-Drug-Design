#!/usr/bin/env python3
"""
Novel Molecules Extractor
Author: TAQDEES
Description: Extract and analyze only the novel molecules from validation results
"""

import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def extract_novel_molecules():
    """Extract only novel molecules from validation results"""
    
    logger.info("🔍 Extracting Novel Molecules")
    logger.info("=" * 50)
    
    # Load validation results
    validation_file = 'modification_results/molecular_properties_validation_results.csv'
    
    if not os.path.exists(validation_file):
        logger.error(f"❌ Validation results file not found: {validation_file}")
        logger.info("Please run the molecule_validator.py script first!")
        return
    
    try:
        df = pd.read_csv(validation_file)
        logger.info(f"📋 Loaded validation results: {len(df)} molecules")
    except Exception as e:
        logger.error(f"Error loading validation file: {e}")
        return
    
    # Filter for novel molecules only
    novel_mask = df['is_novel'] == True
    novel_df = df[novel_mask].copy()
    
    logger.info(f"🆕 Found {len(novel_df)} novel molecules out of {len(df)} total")
    
    if len(novel_df) == 0:
        logger.warning("❌ No novel molecules found!")
        return
    
    # Sort by QED score (drug-likeness) descending
    novel_df_sorted = novel_df.sort_values('qed', ascending=False).reset_index(drop=True)
    
    # Add ranking
    novel_df_sorted['novelty_rank'] = range(1, len(novel_df_sorted) + 1)
    
    # Create output directory
    output_dir = "novel_molecules_only"
    os.makedirs(output_dir, exist_ok=True)
    
    # Save novel molecules only
    novel_molecules_file = f"{output_dir}/novel_molecules_only.csv"
    novel_df_sorted.to_csv(novel_molecules_file, index=False)
    
    # Create a simplified summary
    summary_columns = [
        'novelty_rank', 'smiles', 'qed', 'molecular_weight', 'logp', 
        'max_similarity_known', 'most_similar_known', 'is_drug_like',
        'lipinski_violations', 'scaffold'
    ]
    
    summary_df = novel_df_sorted[summary_columns].copy()
    summary_file = f"{output_dir}/novel_molecules_summary.csv"
    summary_df.to_csv(summary_file, index=False)
    
    # Create detailed analysis report
    create_novel_analysis_report(novel_df_sorted, output_dir)
    
    # Display summary
    logger.info(f"\n🏆 NOVEL MOLECULES RANKING (by QED score):")
    logger.info("=" * 80)
    
    for idx, row in novel_df_sorted.iterrows():
        logger.info(f"Rank {row['novelty_rank']:2d}: QED {row['qed']:.3f} | Similarity {row['max_similarity_known']:.3f}")
        logger.info(f"        {row['smiles']}")
        logger.info(f"        MW: {row['molecular_weight']:.0f} | LogP: {row['logp']:.2f} | Similar to: {row['most_similar_known']}")
        logger.info("")
    
    logger.info(f"💾 Files created in {output_dir}/:")
    logger.info(f"   📄 novel_molecules_only.csv - Complete data for novel molecules")
    logger.info(f"   📄 novel_molecules_summary.csv - Key properties summary")
    logger.info(f"   📄 novel_molecules_analysis.txt - Detailed analysis report")
    
    return novel_df_sorted

def create_novel_analysis_report(novel_df, output_dir):
    """Create detailed analysis report for novel molecules"""
    
    report_file = f"{output_dir}/novel_molecules_analysis.txt"
    
    with open(report_file, 'w') as f:
        f.write("NOVEL EGFR INHIBITOR CANDIDATES - DETAILED ANALYSIS\n")
        f.write("=" * 60 + "\n")
        f.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total Novel Molecules: {len(novel_df)}\n\n")
        
        # Overall statistics
        f.write("SUMMARY STATISTICS:\n")
        f.write("-" * 30 + "\n")
        f.write(f"Average QED Score: {novel_df['qed'].mean():.3f}\n")
        f.write(f"Average Molecular Weight: {novel_df['molecular_weight'].mean():.1f} Da\n")
        f.write(f"Average LogP: {novel_df['logp'].mean():.2f}\n")
        f.write(f"Average Similarity to Known: {novel_df['max_similarity_known'].mean():.3f}\n")
        f.write(f"Drug-like molecules: {novel_df['is_drug_like'].sum()}/{len(novel_df)}\n")
        f.write(f"Lipinski compliant: {(novel_df['lipinski_violations'] == 0).sum()}/{len(novel_df)}\n\n")
        
        # Categories
        high_qed = novel_df[novel_df['qed'] >= 0.7]
        medium_qed = novel_df[(novel_df['qed'] >= 0.5) & (novel_df['qed'] < 0.7)]
        
        f.write("DRUG-LIKENESS CATEGORIES:\n")
        f.write("-" * 30 + "\n")
        f.write(f"High QED (≥0.7): {len(high_qed)} molecules\n")
        f.write(f"Medium QED (0.5-0.7): {len(medium_qed)} molecules\n\n")
        
        # Novelty categories
        very_novel = novel_df[novel_df['max_similarity_known'] < 0.5]
        moderately_novel = novel_df[(novel_df['max_similarity_known'] >= 0.5) & (novel_df['max_similarity_known'] < 0.7)]
        
        f.write("NOVELTY CATEGORIES:\n")
        f.write("-" * 30 + "\n")
        f.write(f"Very Novel (<50% similarity): {len(very_novel)} molecules\n")
        f.write(f"Moderately Novel (50-70% similarity): {len(moderately_novel)} molecules\n\n")
        
        # Detailed molecule analysis
        f.write("DETAILED MOLECULE ANALYSIS:\n")
        f.write("=" * 60 + "\n\n")
        
        for idx, row in novel_df.iterrows():
            f.write(f"RANK {row['novelty_rank']}: {row['smiles']}\n")
            f.write("-" * 60 + "\n")
            f.write(f"Drug-likeness (QED): {row['qed']:.3f}\n")
            f.write(f"Molecular Weight: {row['molecular_weight']:.1f} Da\n")
            f.write(f"LogP: {row['logp']:.2f}\n")
            f.write(f"H-bond Donors: {row['hbd']}\n")
            f.write(f"H-bond Acceptors: {row['hba']}\n")
            f.write(f"TPSA: {row['tpsa']:.1f} Ų\n")
            f.write(f"Rotatable Bonds: {row['rotatable_bonds']}\n")
            f.write(f"Aromatic Rings: {row['aromatic_rings']}\n")
            f.write(f"Lipinski Violations: {row['lipinski_violations']}\n")
            f.write(f"Novelty Score: {row['max_similarity_known']:.3f} (vs {row['most_similar_known']})\n")
            f.write(f"Scaffold: {row['scaffold']}\n")
            
            # Classification
            if row['qed'] >= 0.7:
                qed_class = "HIGH drug-likeness"
            elif row['qed'] >= 0.5:
                qed_class = "MEDIUM drug-likeness"
            else:
                qed_class = "LOW drug-likeness"
            
            if row['max_similarity_known'] < 0.5:
                novelty_class = "VERY NOVEL"
            elif row['max_similarity_known'] < 0.7:
                novelty_class = "MODERATELY NOVEL"
            else:
                novelty_class = "SLIGHTLY NOVEL"
            
            f.write(f"Classification: {qed_class}, {novelty_class}\n")
            
            # Recommendations
            f.write("Recommendations:\n")
            if row['qed'] >= 0.7 and row['max_similarity_known'] < 0.6:
                f.write("  ⭐ HIGH PRIORITY - Excellent drug-likeness and high novelty\n")
                f.write("  🧪 Recommended for experimental synthesis and testing\n")
            elif row['qed'] >= 0.6:
                f.write("  ✅ MEDIUM PRIORITY - Good drug-likeness\n")
                f.write("  🔬 Consider for computational docking studies first\n")
            else:
                f.write("  ⚠️  LOW PRIORITY - May need optimization\n")
                f.write("  📚 Use as starting point for further modifications\n")
            
            f.write("\n")
    
    logger.info(f"📄 Detailed analysis report created: {report_file}")

def create_excel_workbook():
    """Create Excel workbook with multiple sheets for novel molecules"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Load data
        novel_df = pd.read_csv('novel_molecules_only/novel_molecules_only.csv')
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Novel Molecules Summary")
        summary_cols = ['novelty_rank', 'smiles', 'qed', 'molecular_weight', 'logp', 
                       'max_similarity_known', 'most_similar_known']
        summary_data = novel_df[summary_cols]
        
        for r in dataframe_to_rows(summary_data, index=False, header=True):
            ws_summary.append(r)
        
        # Sheet 2: Full Data
        ws_full = wb.create_sheet("Complete Data")
        for r in dataframe_to_rows(novel_df, index=False, header=True):
            ws_full.append(r)
        
        # Sheet 3: High Priority (QED >= 0.7)
        ws_priority = wb.create_sheet("High Priority")
        priority_data = novel_df[novel_df['qed'] >= 0.7]
        for r in dataframe_to_rows(priority_data, index=False, header=True):
            ws_priority.append(r)
        
        # Style headers
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for ws in [ws_summary, ws_full, ws_priority]:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        # Save workbook
        excel_file = "novel_molecules_only/novel_molecules_analysis.xlsx"
        wb.save(excel_file)
        logger.info(f"📊 Excel workbook created: {excel_file}")
        
    except ImportError:
        logger.info("📊 openpyxl not available. Install with: pip install openpyxl")
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")

def main():
    """Main function"""
    
    logger.info("🧬 Novel Molecules Extractor")
    logger.info("Extracting only the novel molecules for focused analysis")
    logger.info("=" * 60)
    
    # Extract novel molecules
    novel_df = extract_novel_molecules()
    
    if novel_df is not None:
        # Create Excel workbook if possible
        create_excel_workbook()
        
        logger.info("\n🎉 Novel molecules extraction complete!")
        logger.info("\n📁 Check the 'novel_molecules_only/' folder for:")
        logger.info("   📄 novel_molecules_only.csv - Complete novel molecules data")
        logger.info("   📄 novel_molecules_summary.csv - Key properties only")
        logger.info("   📄 novel_molecules_analysis.txt - Detailed analysis report")
        logger.info("   📊 novel_molecules_analysis.xlsx - Excel workbook (if available)")
        
        logger.info(f"\n🏆 RECOMMENDATION:")
        logger.info("Focus on the top 3-5 molecules with highest QED scores")
        logger.info("These represent the best balance of novelty and drug-likeness!")

if __name__ == "__main__":
    main()